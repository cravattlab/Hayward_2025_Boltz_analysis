'''Helper module to encode formatting in final Extended Data 2 dataset'''
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font

# Colors
RED = '#fbb4ae'
BLUE = '#b3cde3'
GREEN = '#ccebc5'
PURPLE = '#decbe4'
ORANGE = '#fed9a6'
YELLOW = '#ffffcc'
BEIGE = '#e5d8bd'
PINK = '#fddaec'
GRAY = '#f2f2f2'

# simplified stereoprobe names to match manuscript
LIGAND_MAPPINGS = {
    'MY-5A': 'EV-98',
    'MY-7A': 'EV-96',
    'MY-5B': 'EV-99',
    'MY-7B': 'EV-97',
    'WX-01-09': 'WX-01-09',
    'WX-01-10': 'WX-01-10',
    'WX-01-11': 'WX-01-11',
    'WX-01-12': 'WX-01-12',
    'WX-02-418': 'WX-61a',
    'WX-02-419': 'WX-61b',
    'WX-02-420': 'WX-61c',
    'WX-02-421': 'WX-61d',
    'WX-02-422': 'WX-62a',
    'WX-02-423': 'WX-62b',
    'WX-02-424': 'WX-62c',
    'WX-02-425': 'WX-62d',
    'WX-02-426': 'WX-63a',
    'WX-02-428': 'WX-63b',
    'WX-02-427': 'WX-63c',
    'WX-02-429': 'WX-63d',
    'WX-02-687': 'WX-64a',
    'WX-02-689': 'WX-64b',
    'WX-02-688': 'WX-64c',
    'WX-02-690': 'WX-64d',
    'WX-02-430': 'WX-71a',
    'WX-02-432': 'WX-71b',
    'WX-02-431': 'WX-71c',
    'WX-02-433': 'WX-71d',
    'WX-02-434': 'WX-72a',
    'WX-02-436': 'WX-72b',
    'WX-02-435': 'WX-72c',
    'WX-02-437': 'WX-72d',
    'WX-02-438': 'WX-73a',
    'WX-02-440': 'WX-73b',
    'WX-02-439': 'WX-73c',
    'WX-02-441': 'WX-73d',
    'WX-02-691': 'WX-74a',
    'WX-02-693': 'WX-74b',
    'WX-02-692': 'WX-74c',
    'WX-02-694': 'WX-74d',
    'WX-02-891': 'WX-72a-yne',
    'WX-02-893': 'WX-72b-yne',
    'WX-02-892': 'WX-72c-yne',
    'WX-02-894': 'WX-72d-yne',
    'WX-02-909': 'WX-62a-yne',
    'WX-02-911': 'WX-62b-yne',
    'WX-02-910': 'WX-62c-yne',
    'WX-02-912': 'WX-62d-yne',
    'WX-02-905': 'WX-73a-yne',
    'WX-02-907': 'WX-73b-yne',
    'WX-02-906': 'WX-73c-yne',
    'WX-02-908': 'WX-73d-yne',
}

# Hardcoded sheet names
SHEET_MAPPINGS = {
    'analysis_data': 'analysis data',
    'per_peptide_summary': 'per-peptide summary',
    'orthosteric_sites': 'orthosteric sites',
    'failed_matches': 'failed matches',
    'models_validation': 'physical validation'
}

# Hardcoded column names/colors
# Columns to be removed from input are optionally commented for reference
COLUMN_MAPPINGS = {
    'analysis data': {
        'accession': ('accession', BLUE),
        'gene_name': ('gene name', BLUE),
        'ligand_name': ('stereoprobe', BLUE),
        'cys_site': ('liganded cysteine', BLUE),
        # 'domain identifier': ('NEW NAME', COLOR),
        'model_rank': ('model rank', RED),
        # 'residue in range': ('NEW NAME', COLOR),
        # 'cys in range': ('NEW NAME', COLOR),
        'mean_global_pLDDT': ('mean global pLDDT', RED),
        'stdev_global_pLDDT': ('stdev global pLDDT', RED),
        'mean_ligand_pLDDT': ('mean ligand pLDDT', RED),
        # 'match with cys binding site': ('NEW NAME', COLOR),
        'mean_binding_site_plddt': ('mean pLDDT for residues within 5 Å of stereoprobe', RED),
        'mean_cys_sites_plddt': ('mean pLDDT for residues within 5 Å of cysteine', RED),
        'distance_to_cys': ('stereoprobe–cysteine distance', RED),
        'distance_to_nearest_orthosteric_site': ('stereoprobe–nearest orthosteric site distance', GREEN),
        'nearest_orthosteric_ligand': ('nearest orthosteric ligand name', GREEN),
        'cys_to_nearest_orthosteric_distance': ('cysteine—nearest orthosteric site distance', GREEN),
        'cysteine_orthostery_classification': ('orthosteric classification for cysteine', GREEN),
        'ligand_orthostery_classification': ('orthosteric classification for stereoprobe', GREEN),
        # 'scenario': ('orthosteric classification', GREEN),
        # 'nearest orthosteric from cys': ('NEW NAME', COLOR),
        # 'stereoisomer index': ('NEW NAME', COLOR),
        # 'stereoisomer source': ('NEW NAME', COLOR),
        # 'source smiles': ('NEW NAME', COLOR),
        # 'template SMILES': ('NEW NAME', COLOR),
        # 'predicted ligand SMILES': ('NEW NAME', COLOR),
        'template_stereochemistry': ('stereoprobe stereochemistry', YELLOW),
        'ligand_stereochemistry': ('predicted stereoprobe stereochemistry', YELLOW),
        'stereochemistry_result': ('known vs predicted stereochemistry', YELLOW),
        # 'is stereochemistry match': ('NEW NAME', COLOR),
        # 'has orthosteric distance': ('NEW NAME', COLOR),
        # 'ligand near cys': ('NEW NAME', COLOR),
        # 'ligand near orthosteric': ('NEW NAME', COLOR),
        # 'cys near orthosteric': ('NEW NAME', COLOR),
        # 'Ligand Bridging': ('NEW NAME', COLOR),
        # 'Direct Proximity': ('NEW NAME', COLOR),
        # 'Edge (10 Å)': ('NEW NAME', COLOR),
        # 'Distant': ('NEW NAME', COLOR),
        # 'orthosteric event': ('NEW NAME', COLOR),
        # 'non-orthosteric event': ('NEW NAME', COLOR),
        # 'other': ('NEW NAME', COLOR),
        # 'Correct (no orthosteric annotation)': ('NEW NAME', COLOR),
        # 'Incorrect (no orthosteric annotation)': ('NEW NAME', COLOR),
    },
    'per-peptide summary': {
        'accession': ('accession', BLUE),
        'gene_name': ('gene name', BLUE),
        'ligand_name': ('stereoprobe', BLUE),
        'cys_site': ('liganded cysteine', BLUE),
        'site_group': ('cysteines on tryptic peptide', BLUE),
        # 'fraction near cys': ('NEW NAME', RED),
        # 'domain identifier': ('NEW NAME', COLOR),
        # 'complex name': ('NEW NAME', COLOR),
        # 'model rank': ('model rank', RED),
        # 'model file name': ('NEW NAME', COLOR),
        # 'folder name': ('NEW NAME', COLOR),
        # 'residue ranges': ('NEW NAME', COLOR),
        # 'cys in range': ('NEW NAME', COLOR),
        # 'mean global pLDDT': ('NEW NAME', COLOR),
        # 'stdev global pLDDT': ('NEW NAME', COLOR),
        # 'mean ligand pLDDT': ('NEW NAME', COLOR),
        # 'calc complete': ('NEW NAME', COLOR),
        # 'calc resnos': ('NEW NAME', COLOR),
        # 'match with cys binding site': ('NEW NAME', COLOR),
        # 'resnos around cys site': ('NEW NAME', COLOR),
        # 'mean binding site plddt': ('NEW NAME', COLOR),
        # 'mean cys sites plddt': ('NEW NAME', COLOR),
        # 'plddt scores for each residue around binding site': ('NEW NAME', COLOR),
        # 'plddt scores for each residues around cys site': ('NEW NAME', COLOR),
        'mean_binding_site_plddt': ('mean pLDDT for residues within 5 Å of ligand', RED),
        'mean_cys_sites_plddt': ('mean pLDDT for residues within 5 Å of cysteine', RED),
        'distance_to_cys': ('stereoprobe–cysteine distance', RED),
        'distance_to_nearest_orthosteric_site': ('stereoprobe–nearest orthosteric site distance', GREEN),
        'nearest_orthosteric_ligand': ('nearest orthosteric ligand name', GREEN),
        'cys_to_nearest_orthosteric_distance': ('cysteine–nearest orthosteric site distance', GREEN),
        'cysteine_orthostery_classification': ('orthosteric classification for cysteine', GREEN),
        'ligand_orthostery_classification': ('orthosteric classification for stereoprobe', GREEN),
        # 'stereoisomer index': ('NEW NAME', COLOR),
        # 'stereoisomer source': ('NEW NAME', COLOR),
        # 'source smiles': ('NEW NAME', COLOR),
        # 'template SMILES': ('NEW NAME', COLOR),
        # 'predicted ligand SMILES': ('NEW NAME', COLOR),
        # 'template stereochemistry': ('NEW NAME', COLOR),
        # 'ligand stereochemistry': ('NEW NAME', COLOR),
        # 'stereochemistry result': ('NEW NAME', COLOR),
        # 'is stereochemistry match': ('NEW NAME', COLOR),
        # 'has orthosteric distance': ('NEW NAME', COLOR),
        # 'ligand near cys': ('NEW NAME', COLOR),
        # 'ligand near orthosteric': ('NEW NAME', COLOR),
        # 'cys near orthosteric': ('NEW NAME', COLOR),
        # 'Ligand Bridging': ('NEW NAME', COLOR),
        # 'Direct Proximity': ('NEW NAME', COLOR),
        # 'Edge (10 Å)': ('NEW NAME', COLOR),
        # 'Distant': ('NEW NAME', COLOR),
        # 'orthosteric event': ('NEW NAME', COLOR),
        # 'non-orthosteric event': ('NEW NAME', COLOR),
        # 'other': ('NEW NAME', COLOR),
        # 'Correct (no orthosteric annotation)': ('NEW NAME', COLOR),
        # 'Incorrect (no orthosteric annotation)': ('NEW NAME', COLOR),
        # 'scenario': ('NEW NAME', COLOR),
    },
    'orthosteric sites': {
        'accession': ('accession', BLUE),
        'gene_name': ('gene name', BLUE),
        'orthosteric_ligand_name': ('orthosteric ligand name', GREEN),
        'orthosteric_ligand_binding_site': ('annotated residue', GREEN),
        'orthosteric_ligand_note': ('note', GREEN),
        'description': ('description', GREEN),
        'feature_type': ('feature type', GREEN),
        'status': ('status', GREEN),
    },
    'failed matches': {
        'accession': ('accession', BLUE),
        'gene_name': ('gene name', BLUE),
        # 'protein name': ('NEW NAME', COLOR),
        'ligand_name': ('stereoprobe', BLUE),
        'cys_site': ('liganded cysteine', BLUE),
        # 'reason': ('NEW NAME', COLOR),
        # 'Protein': ('NEW NAME', COLOR),
        # 'Site': ('NEW NAME', COLOR),
        # 'Accession': ('NEW NAME', COLOR),
        # 'Description': ('NEW NAME', COLOR),
        # 'Number of observations (biological replicates)': ('NEW NAME', COLOR),
        # 'IA-DTB blockade variability (%)': ('NEW NAME', COLOR),
        # 'IA-DTB blockade average (%)': ('NEW NAME', COLOR),
        # 'Fold-enantioselectivity': ('NEW NAME', COLOR),
        # 'Diastereoselectivity': ('NEW NAME', COLOR),
        # 'Chemoselectivity': ('NEW NAME', COLOR),
        # 'Stereoprobe': ('NEW NAME', COLOR),
        # 'Stereoconfiguration': ('NEW NAME', COLOR),
        # 'Experiment	Site-specificity': ('NEW NAME', COLOR),
        # 'Uniprot ID': ('NEW NAME', COLOR),
        # 'Evidence of stereoselective enrichment (Njomen et al)': ('NEW NAME', COLOR),
        # 'No site spec and no stereo enrichment': ('NEW NAME', COLOR),
    },
    'physical validation': {
        'accession': ('accession', BLUE),
        'gene_name': ('gene name', BLUE),
        'ligand_name': ('stereoprobe', BLUE),
        'cys_site': ('liganded cysteine', BLUE),
        'model_rank': ('model rank', RED),
        # 'mol pred loaded': ('NEW NAME', COLOR),
        # 'mol true loaded': ('NEW NAME', COLOR),
        # 'mol cond loaded': ('NEW NAME', COLOR),
        # 'sanitization': ('NEW NAME', COLOR),
        # 'inchi convertible': ('NEW NAME', COLOR),
        # 'all atoms connected': ('NEW NAME', COLOR),
        # 'molecular formula': ('NEW NAME', COLOR),
        # 'molecular bonds': ('NEW NAME', COLOR),
        # 'double bond stereochemistry': ('NEW NAME', COLOR),
        # 'tetrahedral chirality': ('NEW NAME', COLOR),
        'bond_lengths': ('bond lengths', PINK),
        'bond_angles': ('bond angles', PINK),
        'internal_steric_clash': ('internal steric clash', PINK),
        # 'aromatic_ring_flatness': ('aromatic ring flatness', PINK),
        # 'non-aromatic_ring_non-flatness': ('non-aromatic ring flatness', PINK),
        # 'double_bond_flatness': ('double bond flatness', PINK),
        'protein-ligand_maximum_distance': ('protein–stereoprobe maximum distance', PINK),
        'minimum_distance_to_protein': ('minimum distance to protein', PINK),
        # 'minimum distance to organic cofactors': ('NEW NAME', COLOR),
        # 'minimum distance to inorganic cofactors': ('NEW NAME', COLOR),
        # 'minimum distance to waters': ('NEW NAME', COLOR),
        'volume_overlap_with_protein': ('volume overlap with protein', PINK),
        # 'volume overlap with organic cofactors': ('NEW NAME', COLOR),
        # 'volume overlap with inorganic cofactors': ('NEW NAME', COLOR),
        # 'volume overlap with waters': ('NEW NAME', COLOR),
        # 'rmsd ≤ 2å': ('NEW NAME', COLOR),
        # 'passes valence checks': ('NEW NAME', PINK),
        # 'passes kekulization': ('NEW NAME', COLOR),
        # 'inchi crystal valid': ('NEW NAME', COLOR),
        # 'inchi docked valid': ('NEW NAME', COLOR),
        # 'inchi crystal': ('NEW NAME', COLOR),
        # 'inchi docked': ('NEW NAME', COLOR),
        # 'number bonds': ('NEW NAME', COLOR),
        'shortest_bond_relative_length': ('shortest bond relative length', PINK),
        'longest_bond_relative_length': ('longest bond relative length', PINK),
        'number_short_outlier_bonds': ('number short outlier bonds', PINK),
        'number_long_outlier_bonds': ('number long outlier bonds', PINK),
        'number_angles': ('number angles', PINK),
        'most_extreme_relative_angle': ('most extreme relative angle', PINK),
        # 'number outlier angles': ('number outlier angles', PINK),
        # 'number noncov pairs': ('NEW NAME', COLOR),
        'shortest_noncovalent_relative_distance': ('shortest noncovalent relative distance', PINK),
        'number_clashes': ('number clashes', PINK),
        'number_valid_bonds': ('number valid bonds', PINK),
        'number_valid_angles': ('number valid angles', PINK),
        'number_valid_noncov_pairs': ('number valid noncov pairs', PINK),
        # 'number aromatic rings checked': ('NEW NAME', COLOR),
        # 'number aromatic rings pass': ('NEW NAME', COLOR),
        # 'aromatic ring maximum distance from plane': ('NEW NAME', COLOR),
        # 'number non-aromatic rings checked': ('NEW NAME', COLOR),
        # 'number non-aromatic rings pass': ('NEW NAME', COLOR),
        # 'non-aromatic ring maximum distance from plane': ('NEW NAME', COLOR),
        # 'number double bonds checked': ('NEW NAME', COLOR),
        # 'number double bonds pass': ('NEW NAME', COLOR),
        # 'double bond maximum distance from plane': ('NEW NAME', COLOR),
        'smallest_distance_protein': ('smallest distance protein', PINK),
        'num_pairwise_clashes_protein': ('num pairwise clashes protein', PINK),
        # 'most extreme ligand atom id protein': ('NEW NAME', COLOR),
        # 'most extreme protein atom id protein': ('NEW NAME', COLOR),
        # 'most extreme ligand element protein': ('NEW NAME', COLOR),
        # 'most extreme protein element protein': ('NEW NAME', COLOR),
        'most_extreme_ligand_vdw_protein': ('most extreme stereoprobe vdw protein', PINK),
        'most_extreme_protein_vdw_protein': ('most extreme protein vdw protein', PINK),
        'most_extreme_sum_radii_protein': ('most extreme sum radii protein', PINK),
        'most_extreme_distance_protein': ('most extreme distance protein', PINK),
        'most_extreme_sum_radii_scaled_protein': ('most extreme sum radii scaled protein', PINK),
        'most_extreme_relative_distance_protein': ('most extreme relative distance protein', PINK),
        # 'most extreme clash protein': ('NEW NAME', COLOR),
        # 'smallest distance organic cofactors': ('NEW NAME', COLOR),
        # 'not too far away organic cofactors': ('NEW NAME', COLOR),
        # 'num pairwise clashes organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme ligand atom id organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme protein atom id organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme ligand element organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme protein element organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme ligand vdw organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme protein vdw organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme sum radii organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme distance organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme sum radii scaled organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme relative distance organic cofactors': ('NEW NAME', COLOR),
        # 'most extreme clash organic cofactors': ('NEW NAME', COLOR),
        # 'smallest distance inorganic cofactors': ('NEW NAME', COLOR),
        # 'not too far away inorganic cofactors': ('NEW NAME', COLOR),
        # 'num pairwise clashes inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme ligand atom id inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme protein atom id inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme ligand element inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme protein element inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme ligand vdw inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme protein vdw inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme sum radii inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme distance inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme sum radii scaled inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme relative distance inorganic cofactors': ('NEW NAME', COLOR),
        # 'most extreme clash inorganic cofactors': ('NEW NAME', COLOR),
        # 'smallest distance waters': ('NEW NAME', COLOR),
        # 'not too far away waters': ('NEW NAME', COLOR),
        # 'num pairwise clashes waters': ('NEW NAME', COLOR),
        # 'most extreme ligand atom id waters': ('NEW NAME', COLOR),
        # 'most extreme protein atom id waters': ('NEW NAME', COLOR),
        # 'most extreme ligand element waters': ('NEW NAME', COLOR),
        # 'most extreme protein element waters': ('NEW NAME', COLOR),
        # 'most extreme ligand vdw waters': ('NEW NAME', COLOR),
        # 'most extreme protein vdw waters': ('NEW NAME', COLOR),
        # 'most extreme sum radii waters': ('NEW NAME', COLOR),
        # 'most extreme distance waters': ('NEW NAME', COLOR),
        # 'most extreme sum radii scaled waters': ('NEW NAME', COLOR),
        # 'most extreme relative distance waters': ('NEW NAME', COLOR),
        # 'most extreme clash waters': ('NEW NAME', COLOR),
        # 'volume overlap protein': ('NEW NAME', COLOR),
        # 'volume overlap organic cofactors': ('NEW NAME', COLOR),
        # 'volume overlap inorganic cofactors': ('NEW NAME', COLOR),
        # 'volume overlap waters': ('NEW NAME', COLOR),
        # 'rmsd': ('NEW NAME', COLOR),
        # 'kabsch rmsd': ('NEW NAME', COLOR),
        # 'centroid distance': ('NEW NAME', COLOR),
        # 'phenix available': ('NEW NAME', COLOR),
        # 'phenix error': ('NEW NAME', COLOR),
        'bond_mean': ('mean bond deviation from ideal (Å)', BEIGE),
        'bond_max': ('max bond deviation from ideal (Å)', BEIGE),
        'bond_rmsz': ('bond deviation rmsZ', BEIGE),
        'angle_mean': ('mean angle deviation from ideal (°)', BEIGE),
        'angle_max': ('max angle deviation from ideal (°)', BEIGE),
        'angle_rmsz': ('angle deviation rmsZ', BEIGE),
        'chirality_mean': ('mean chirality', BEIGE),
        'planarity_mean': ('mean planarity', BEIGE),
        'dihedral_mean': ('mean dihedral', BEIGE),
        'min_nonbonded_distance': ('min nonbonded distance', BEIGE),
        'clashscore': ('all-atom clashscore', BEIGE),
        'rama_outliers_pct': ('Ramachandran outliers (%)', BEIGE),
        'rama_allowed_pct': ('Ramachandran allowed (%)', BEIGE),
        'rama_favored_pct': ('Ramachandran favored (%)', BEIGE),
        'rotamer_outliers_pct': ('rotamer outliers (%)', BEIGE),
        'rotamer_allowed_pct': ('rotamer allowed (%)', BEIGE),
        'rotamer_favored_pct': ('rotamer favored (%)', BEIGE),
        # 'cbeta deviations pct': ('NEW NAME', COLOR),
        # 'inchi overall': ('NEW NAME', COLOR),
        # 'inchi version': ('NEW NAME', COLOR),
        # 'hydrogens': ('NEW NAME', COLOR),
        # 'net charge': ('NEW NAME', COLOR),
        # 'protons': ('NEW NAME', COLOR),
        # 'stereo sp3': ('NEW NAME', COLOR),
        # 'stereo sp3 inverted': ('NEW NAME', COLOR),
        # 'stereo type': ('NEW NAME', COLOR),
        # 'stereochemistry preserved': ('NEW NAME', COLOR),
        # 'pb error': ('NEW NAME', COLOR),
        # 'pb valid': ('NEW NAME', COLOR),
    }
}


def remap_ligand_names(df: pd.DataFrame) -> pd.DataFrame:
    """Remap ligand names in stereoprobe column if it exists."""
    if 'stereoprobe' in df.columns:
        df['stereoprobe'] = df['stereoprobe'].map(
            lambda x: LIGAND_MAPPINGS.get(x, x)
        )
    return df


def save_formatted_excel(
    output_file: str | Path,
    dataframes: dict[str, pd.DataFrame],
    save_raw_data: bool = False,
    remap_ligands: bool = True
) -> None:
    """Save dataframes to Excel with formatted headers and styling."""
    output_path = Path(output_file)

    # Save raw data as separate file if requested
    if save_raw_data:
        raw_path = output_path.with_stem(f"{output_path.stem}_raw")
        with pd.ExcelWriter(raw_path, engine='openpyxl') as raw_writer:
            for sheet_name, df in dataframes.items():
                df.to_excel(raw_writer, sheet_name=sheet_name, index=False)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        sheets_written = 0

        for sheet_name, df in dataframes.items():
            column_colors = {}

            # update sheet names first for consistency
            display_name = SHEET_MAPPINGS.get(sheet_name, sheet_name)

            df_to_save = df.copy()

            # only include mapped sheets in formatted Excel file
            if display_name in COLUMN_MAPPINGS:
                mapping = COLUMN_MAPPINGS[display_name]

                # Parse mapping to rename + color
                parsed_mapping = {}
                for col, value in mapping.items():
                    if isinstance(value, tuple):
                        parsed_mapping[col] = value[0]
                        column_colors[value[0]] = value[1]
                    else:
                        parsed_mapping[col] = value

                missing_cols = [c for c in parsed_mapping if c not in df.columns]
                cols_to_keep = list(parsed_mapping.keys())

                if missing_cols:
                    if df.empty:
                        # Populate an empty sheet with the expected headers (e.g. when validation is skipped)
                        df_filtered = pd.DataFrame(columns=cols_to_keep)
                    else:
                        raise KeyError(f"Required columns not found in {display_name}: {missing_cols}")
                else:
                    df_filtered = df[cols_to_keep].copy()

                df_to_save = df_filtered.rename(columns={c: parsed_mapping[c] for c in cols_to_keep})

                if remap_ligands:
                    df_to_save = remap_ligand_names(df_to_save)

            # Write and format
            df_to_save.to_excel(writer, sheet_name=display_name, index=False)
            sheets_written += 1

            ws = writer.sheets[display_name]
            for idx, column in enumerate(ws.columns):
                max_length = max((len(str(c.value)) for c in column if c.value), default=0)
                col_letter = column[0].column_letter
                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)

                header_cell = ws.cell(row=1, column=idx + 1)
                if header_cell.value in column_colors:
                    color_hex = column_colors[header_cell.value].replace('#', '')
                    header_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                header_cell.font = Font(bold=True)
